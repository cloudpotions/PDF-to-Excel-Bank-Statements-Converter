#!/usr/bin/env python3
"""
Chase Bank Statement Processor
Extracts transactions from Chase PDF statements and organizes them into Excel.

Usage:
    python3 pdf-2-excel.py                 # opens a folder picker
    python3 pdf-2-excel.py /path/to/folder # process a folder directly (no dialog)
    python3 pdf-2-excel.py --help          # show help

Notes:
  * Pass the folder path as an argument to skip the picker entirely - this is the
    fastest, least clunky way, especially on Linux.
  * On Linux, the picker uses PyGObject (Gtk.FileChooserNative), which on a
    Wayland session routes through xdg-desktop-portal and shows the modern
    native dialog (sidebar + large icons). Both python3-gi and gir1.2-gtk-3.0
    are preinstalled on Ubuntu desktop, so there is nothing to install. If
    PyGObject is missing, the script falls back to `zenity`, then Tk.
  * Works headless (over SSH / no display) as long as you pass the folder path.
"""
import os
import re
import sys
import shutil
import platform
import subprocess
import traceback
from datetime import datetime

HELP = __doc__


# ---------------------------------------------------------------------------
# Cross-platform, low-friction user I/O
# ---------------------------------------------------------------------------
class UI:
    """Talk to the user via Tk dialogs when a display is available, otherwise
    fall back to plain console text. This keeps the tool usable everywhere -
    desktop GUI, double-click, terminal, or headless over SSH."""

    def __init__(self):
        self.tk = None
        self.root = None
        # Only attempt a GUI if there is a display (or we are on Windows/macOS).
        has_display = (
            platform.system() in ("Windows", "Darwin")
            or os.environ.get("DISPLAY")
            or os.environ.get("WAYLAND_DISPLAY")
        )
        if has_display:
            try:
                import tkinter as tk
                from tkinter import messagebox, filedialog  # noqa: F401
                self.tk = tk
                self.root = tk.Tk()
                self.root.withdraw()  # hide the empty root window
            except Exception:
                self.tk = None
                self.root = None

    @property
    def gui(self):
        return self.root is not None

    def info(self, title, message):
        if self.gui:
            from tkinter import messagebox
            messagebox.showinfo(title, message)
        else:
            print(f"\n=== {title} ===\n{message}\n")

    def welcome(self, title, message, link_text, url):
        """Like info(), but with a clickable link (to the adapt-it-for-another-bank
        guide). Falls back to a plain message + the URL as text if anything fails or
        there's no display."""
        if not self.gui:
            print(f"\n=== {title} ===\n{message}\n\n{link_text}\n{url}\n")
            return
        try:
            import tkinter as tk
            import webbrowser
            top = tk.Toplevel(self.root)
            top.title(title)
            top.resizable(False, False)
            frm = tk.Frame(top, padx=22, pady=18)
            frm.pack(fill="both", expand=True)
            tk.Label(frm, text=message, justify="left", wraplength=460).pack(anchor="w")
            link = tk.Label(frm, text=link_text, fg="#1a4fa0", cursor="hand2",
                            font=("TkDefaultFont", 10, "underline"))
            link.pack(anchor="w", pady=(10, 0))
            link.bind("<Button-1>", lambda e: webbrowser.open(url))
            tk.Button(frm, text="OK", width=12, command=top.destroy).pack(pady=(16, 0))
            top.update_idletasks()
            w, h = top.winfo_width(), top.winfo_height()
            x = (top.winfo_screenwidth() - w) // 2
            y = (top.winfo_screenheight() - h) // 3
            top.geometry(f"+{x}+{y}")
            top.grab_set()
            self.root.wait_window(top)
        except Exception:
            self.info(title, f"{message}\n\n{link_text}\n{url}")

    def error(self, title, message):
        if self.gui:
            from tkinter import messagebox
            messagebox.showerror(title, message)
        else:
            print(f"\n!!! {title} !!!\n{message}\n", file=sys.stderr)

    def confirm(self, title, message):
        if self.gui:
            from tkinter import messagebox
            return messagebox.askyesno(title, message)
        try:
            return input(f"\n{title}\n{message} [Y/n] ").strip().lower() in ("", "y", "yes")
        except EOFError:
            return True


# ---------------------------------------------------------------------------
# Folder selection - the part that used to feel clunky on Linux
# ---------------------------------------------------------------------------
# Sentinel returned by a picker when that picker is *not installed* on this
# system, so the caller can try the next one in the chain. Distinct from None,
# which means "user cancelled" - in that case we DO NOT keep trying other
# pickers (that's what produced the double-dialog bug).
_PICKER_UNAVAILABLE = object()


def default_start_dir():
    """A sensible place to open the picker so the user is not dumped in '/'."""
    home = os.path.expanduser("~")
    for sub in ("Desktop", "Documents", "Downloads"):
        candidate = os.path.join(home, sub)
        if os.path.isdir(candidate):
            return candidate
    return home


def pick_folder_gtk(start_dir):
    """Native GTK folder picker via PyGObject. On Wayland this routes through
    xdg-desktop-portal and produces the modern dialog (sidebar + large icons)."""
    if platform.system() != "Linux":
        return _PICKER_UNAVAILABLE
    try:
        import gi
        gi.require_version("Gtk", "3.0")
        from gi.repository import Gtk
    except (ImportError, ValueError):
        return _PICKER_UNAVAILABLE

    dialog = Gtk.FileChooserNative.new(
        "Select the folder containing your Chase PDF statements",
        None,
        Gtk.FileChooserAction.SELECT_FOLDER,
        "_Open",
        "_Cancel",
    )
    if start_dir and os.path.isdir(start_dir):
        dialog.set_current_folder(start_dir)

    response = dialog.run()
    path = dialog.get_filename() if response == Gtk.ResponseType.ACCEPT else None
    dialog.destroy()
    # Drain pending events so the dialog actually disappears before we return.
    while Gtk.events_pending():
        Gtk.main_iteration()
    return path


def pick_folder_zenity(start_dir):
    """Fallback for Linux systems without PyGObject. Note: zenity 4.0.x on
    Wayland has been observed to mis-report exit codes, which used to cause us
    to open a Tk dialog on top of it - hence we now prefer GTK directly above."""
    if platform.system() != "Linux" or not shutil.which("zenity"):
        return _PICKER_UNAVAILABLE
    try:
        result = subprocess.run(
            [
                "zenity", "--file-selection", "--directory",
                "--title=Select the folder containing your Chase PDF statements",
                f"--filename={start_dir}/",
            ],
            capture_output=True, text=True, timeout=600,
        )
        if result.returncode == 0:
            return result.stdout.strip() or None
        return None
    except Exception:
        return None


def pick_folder_tk(ui, start_dir):
    """Tk's built-in folder chooser. Last-resort fallback for Linux and the
    primary picker on Windows / macOS (where Tk's native dialog is fine)."""
    if not ui.gui:
        return _PICKER_UNAVAILABLE
    from tkinter import filedialog
    path = filedialog.askdirectory(
        title="Select the folder containing your Chase PDF statements",
        initialdir=start_dir,
        mustexist=True,
    )
    return path or None


def choose_folder(ui):
    """Resolve the target folder: command-line arg first, then native picker.
    Returns an absolute path or None. Critically, once a picker is found to be
    available, we use its result as-is - we do not fall through to a second
    picker just because the user cancelled the first."""
    # 1) Explicit path on the command line - fastest, works headless.
    args = [a for a in sys.argv[1:] if not a.startswith("-")]
    if args:
        path = os.path.abspath(os.path.expanduser(args[0]))
        if not os.path.isdir(path):
            ui.error("Folder Not Found", f"That path is not a folder:\n{path}")
            return None
        return path

    start_dir = default_start_dir()

    if platform.system() == "Linux":
        picker_chain = (
            lambda: pick_folder_gtk(start_dir),
            lambda: pick_folder_zenity(start_dir),
            lambda: pick_folder_tk(ui, start_dir),
        )
    else:
        # On Windows / macOS, Tk's filedialog is the native dialog.
        picker_chain = (lambda: pick_folder_tk(ui, start_dir),)

    for picker in picker_chain:
        result = picker()
        if result is not _PICKER_UNAVAILABLE:
            return result

    ui.error(
        "No Way to Pick a Folder",
        "No graphical folder picker is available on this system.\n\n"
        "Just pass the folder path directly, e.g.:\n"
        "    python3 pdf-2-excel.py \"/home/you/Chase Statements 2024\"",
    )
    return None


# ---------------------------------------------------------------------------
# Extraction logic (unchanged behaviour - parses Chase statement text)
# ---------------------------------------------------------------------------
def extract_statement_date(filename):
    """Extract date from filename for sorting (expects YYYYMMDD somewhere)."""
    date_match = re.search(r'(\d{4})(\d{2})(\d{2})', filename)
    if date_match:
        year, month, day = date_match.groups()
        return f"{year}-{month}-{day}"
    return filename


def determine_transaction_year(transaction_date, statement_date):
    """Determine the correct year for a transaction based on the statement date.
    transaction_date: MM/DD   statement_date: YYYY-MM-DD"""
    trans_month = int(transaction_date.split('/')[0])
    statement_year = int(statement_date.split('-')[0])
    statement_month = int(statement_date.split('-')[1])
    # January statement listing a December transaction -> previous year
    if statement_month == 1 and trans_month == 12:
        return str(statement_year - 1)
    # December statement listing a January transaction -> next year
    elif statement_month == 12 and trans_month == 1:
        return str(statement_year + 1)
    return str(statement_year)


def extract_chase_transactions(pages, filename, statement_date):
    """Extract transactions from a Chase statement.

    pages: a list of (page_number, page_text) tuples in PDF order. We walk the
    pages one at a time (instead of merging them into a single blob) so we can
    record which PDF page each transaction was found on. The
    in_transaction_section flag is kept outside the page loop so it still carries
    across page breaks, exactly as before."""
    transactions = []
    in_transaction_section = False
    for page_number, text in pages:
        for line in text.split('\n'):
            if "TRANSACTION DETAIL" in line:
                in_transaction_section = True
                continue
            if not in_transaction_section:
                continue
            if "Beginning Balance" in line:
                continue

            # At each page break Chase emits a "*start*/*end*transaction detail"
            # marker that the PDF text layer merges onto the front of the adjacent
            # transaction line, and the marker absorbs the first digit of the date:
            #   "*end*transac0tion detail1/05 Zelle Payment ... 575.00 6,263.00"
            # Such a line no longer begins with MM/DD, so the date match below
            # skipped it and the transaction was lost. Strip the marker and put the
            # absorbed date digit back so the line parses normally.
            line = re.sub(r'^\*(?:start|end)\*transac(\d?)tion detail', r'\1', line.strip())

            date_match = re.match(r'^(\d{2}/\d{2})', line.strip())
            if date_match:
                current_date = date_match.group(1)
                amounts = re.findall(r'-?[\d,]+\.\d{2}', line)
                if len(amounts) >= 2:
                    amount_str = amounts[-2]
                    balance_str = amounts[-1]
                    amount = float(amount_str.replace(',', ''))
                    balance = float(balance_str.replace(',', ''))
                    desc_start = line.find(current_date) + len(current_date)
                    desc_end = line.rfind(amount_str)
                    description = line[desc_start:desc_end].strip()

                    transaction_year = determine_transaction_year(current_date, statement_date)
                    full_date = f"{current_date}/{transaction_year}"

                    transactions.append({
                        'Statement_Date': statement_date,
                        'Date': full_date,
                        'Description': description,
                        'Amount': amount,
                        'Balance': balance,
                        'Original_Line': line.strip(),
                        'Source_File': filename,
                        'Source_File_Page_Number': page_number,
                    })
    return transactions


# ---------------------------------------------------------------------------
# Make the output workbook nice to look at
# ---------------------------------------------------------------------------
def prettify_sheet(ws):
    """Style an openpyxl worksheet so it's pleasant out of the box: a colored,
    bold, frozen header row; column widths sized to the content; a filter on the
    header; and money columns formatted with thousands separators (negatives in
    red)."""
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="1F4E78")   # dark blue
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="left", vertical="center")

    headers = [c.value for c in ws[1]]

    # Header row: color + bold white text, and a little taller.
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
    ws.row_dimensions[1].height = 20

    # Freeze the header row and add filter dropdowns.
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    money_cols = {"Amount", "Balance"}
    money_fmt = '#,##0.00;[Red]-#,##0.00'
    width_cap = {"Original_Line": 60, "Description": 45}

    for idx, name in enumerate(headers, start=1):
        letter = get_column_letter(idx)
        longest = len(str(name))
        for cell in ws[letter][1:]:
            if cell.value is not None:
                longest = max(longest, len(str(cell.value)))
        cap = width_cap.get(str(name), 22)
        width = min(max(longest + 2, 10), cap)
        width = max(width, len(str(name)) + 2)   # never cut off the header text
        ws.column_dimensions[letter].width = width
        if name in money_cols:
            for cell in ws[letter][1:]:
                cell.number_format = money_fmt
                cell.alignment = Alignment(horizontal="right")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    if any(a in ("-h", "--help") for a in sys.argv[1:]):
        print(HELP)
        return

    ui = UI()

    # Dependencies
    try:
        import pandas as pd
        import pdfplumber
    except ImportError as e:
        ui.error(
            "Missing Dependencies",
            f"Required libraries not found: {e}\n\n"
            "Install them with:\n"
            "    pip install pdfplumber pandas openpyxl\n\n"
            "On some Linux systems you may need:\n"
            "    pip install --break-system-packages pdfplumber pandas openpyxl",
        )
        return

    if not (sys.argv[1:] and not sys.argv[1].startswith("-")):
        ui.welcome(
            "PDF -> Excel Bank Statement Converter",
            "This tool extracts transactions from Chase PDF statements into Excel.\n\n"
            "It's built and tested for Chase checking statements — but you can easily "
            "tweak it for ANY other bank or credit-card statement using AI (about 5 "
            "minutes, no coding needed). The link below shows you how.\n\n"
            "Next: pick the folder that holds your PDF statements. Every PDF in that "
            "folder will be processed.\n\n"
            "Tip: run  python3 pdf-2-excel.py \"/path/to/folder\"  to skip this picker.",
            "How to use it with another bank or credit card (opens GitHub)",
            "https://github.com/cloudpotions/PDF-to-Excel-Bank-Statements-Converter"
            "#use-it-with-any-bank-or-credit-card-5-minute-ai-tweak",
        )

    # Choose the folder
    folder = choose_folder(ui)
    if not folder:
        ui.info("Cancelled", "No folder was selected. Exiting.")
        return
    print(f"\nSelected folder: {folder}")

    # Find PDFs (in the folder, or fall back to the first subfolder that has them)
    pdf_files = [f for f in os.listdir(folder) if f.lower().endswith('.pdf')]
    if not pdf_files:
        for sub in sorted(os.listdir(folder)):
            sub_path = os.path.join(folder, sub)
            if os.path.isdir(sub_path):
                sub_pdfs = [f for f in os.listdir(sub_path) if f.lower().endswith('.pdf')]
                if sub_pdfs:
                    folder, pdf_files = sub_path, sub_pdfs
                    print(f"Found PDFs in subfolder: {sub}\nUsing folder: {folder}")
                    break

    if not pdf_files:
        ui.error(
            "No PDFs Found",
            f"No PDF files were found in:\n{folder}\n\n"
            "Point the tool at a folder that contains your Chase PDF statements.",
        )
        return

    # Confirm before doing the work
    if not ui.confirm(
        "Ready to Process",
        f"Found {len(pdf_files)} PDF file(s) in:\n{folder}\n\nProcess them now?",
    ):
        ui.info("Cancelled", "No problem - nothing was processed.")
        return

    try:
        # Sort files by statement date encoded in the filename
        dated = sorted(
            ((extract_statement_date(f), f) for f in pdf_files),
            key=lambda x: x[0],
        )

        all_transactions = []
        print("\n1. PROCESSING STATEMENTS")
        print("-" * 30)
        for statement_date, filename in dated:
            pdf_path = os.path.join(folder, filename)
            print(f"\nProcessing: {filename} (Statement Date: {statement_date})")
            try:
                with pdfplumber.open(pdf_path) as pdf:
                    # (page_number, text) with page_number 1-based so it matches
                    # the page you see in a PDF viewer.
                    pages = [(i + 1, page.extract_text() or "") for i, page in enumerate(pdf.pages)]
                transactions = extract_chase_transactions(pages, filename, statement_date)
                print(f"Found {len(transactions)} transactions")
                for i, t in enumerate(transactions):
                    t['Statement_Sequence'] = i + 1
                all_transactions.extend(transactions)
            except Exception as e:
                print(f"Error processing {filename}: {e}")
                continue

        if not all_transactions:
            ui.error("No Transactions", "No transactions were found in the PDF files.")
            return

        df = pd.DataFrame(all_transactions)
        df['Sort_Date'] = pd.to_datetime(df['Date'])
        df = df.sort_values(['Sort_Date', 'Statement_Sequence'])

        folder_name = os.path.basename(os.path.normpath(folder))
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_filename = f"{folder_name}_transactions_{timestamp}.xlsx"
        output_path = os.path.join(folder, excel_filename)

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.drop(['Sort_Date', 'Statement_Sequence'], axis=1).to_excel(
                writer, sheet_name='Transactions', index=False)
            df[['Statement_Date', 'Date', 'Description', 'Amount', 'Original_Line',
                'Source_File', 'Source_File_Page_Number']].to_excel(
                writer, sheet_name='Verification', index=False)
            for ws in writer.sheets.values():
                prettify_sheet(ws)

        # Console report
        print("\n2. VERIFICATION SUMMARY")
        print("-" * 30)
        print(f"Total statements processed: {len(dated)}")
        print(f"Total transactions found: {len(df)}")

        print("\n3. FINANCIAL SUMMARY")
        print("-" * 30)
        credits = df[df['Amount'] > 0]['Amount'].sum()
        debits = df[df['Amount'] < 0]['Amount'].sum()
        print(f"Total Credits: ${credits:,.2f}")
        print(f"Total Debits: ${debits:,.2f}")
        print(f"Net Change: ${(credits + debits):,.2f}")

        print("\n4. MONTHLY SUMMARY")
        print("-" * 30)
        df['Month'] = pd.to_datetime(df['Date']).dt.strftime('%Y-%m')
        print(df.groupby('Month').agg({'Amount': ['count', 'sum']}).round(2))

        print(f"\nResults saved to: {output_path}")

        ui.info(
            "Processing Complete",
            f"Processed {len(dated)} statement(s) and {len(df)} transactions.\n\n"
            f"Saved in the same folder as your PDFs:\n{excel_filename}\n\n"
            "Reminder: always spot-check the Excel against the PDFs.",
        )

    except Exception as e:
        traceback.print_exc()
        ui.error("Error", f"An error occurred:\n{e}\n\nSee the console for details.")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"\nAn error occurred: {e}")
        # Keep a double-clicked window open long enough to read the message.
        if sys.stdin and sys.stdin.isatty():
            input("\nPress Enter to exit...")
